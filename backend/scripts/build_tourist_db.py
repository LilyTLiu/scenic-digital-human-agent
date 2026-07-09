"""
将 140K 游客行为数据导入 SQLite，方便查询分析
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import sqlite3
from openpyxl import load_workbook

DATA_FILE = os.path.join(os.path.dirname(__file__), "..", "..", "示范景区公开资料包", "景点景区旅游数据行为分析数据.xlsx")
DB_PATH = os.path.join(os.path.dirname(__file__), "..", "tourist_behavior.db")

def build():
    wb = load_workbook(DATA_FILE, read_only=True)
    ws = wb.active
    conn = sqlite3.connect(DB_PATH)
    conn.execute("DROP TABLE IF EXISTS tourist_behaviors")
    conn.execute("""
        CREATE TABLE tourist_behaviors (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tourist_id TEXT, nickname TEXT, age INTEGER, gender TEXT,
            attraction_name TEXT, attraction_type TEXT,
            visit_date TEXT, stay_duration REAL,
            ticket_cost REAL, food_cost REAL, shopping_cost REAL,
            transport_cost REAL, entertainment_cost REAL, total_cost REAL,
            group_size INTEGER, satisfaction INTEGER
        )
    """)

    batch = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row[0]: continue
        batch.append((
            str(row[0]), str(row[1]), int(row[2] or 0), str(row[3]),
            str(row[4]), str(row[6]), str(row[7])[:10] if row[7] else "",
            float(row[8] or 0), float(row[9] or 0), float(row[10] or 0),
            float(row[11] or 0), float(row[12] or 0), float(row[13] or 0),
            float(row[14] or 0), int(row[15] or 1), int(row[16] or 3),
        ))
        if len(batch) >= 5000:
            conn.executemany("INSERT INTO tourist_behaviors VALUES (NULL,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", batch)
            batch = []
            print(f"  {i+1} / {ws.max_row}")

    if batch:
        conn.executemany("INSERT INTO tourist_behaviors VALUES (NULL,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", batch)

    conn.commit()
    conn.execute("CREATE INDEX IF NOT EXISTS idx_attraction ON tourist_behaviors(attraction_name)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_satisfaction ON tourist_behaviors(satisfaction)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_age ON tourist_behaviors(age)")
    print(f"Done: {conn.execute('SELECT COUNT(*) FROM tourist_behaviors').fetchone()[0]} rows")
    conn.close()

if __name__ == "__main__":
    build()
